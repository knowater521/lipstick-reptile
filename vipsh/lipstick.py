import collections
from pprint import pprint
from openpyxl import Workbook
from selenium import webdriver


class lipstick(object):
    def __init__(self):
        self.driver = webdriver.Chrome(r'D:\pycharm\chromedriver_win32\chromedriver.exe')
        self.item = {}
        self.litem = {}
        self.wb = Workbook()  # 得到一个sheet 表
        self.ws = self.wb.active  # 获取第一个sheet
        self.title_length = 0
        self.name_length = 0
        self.img_length = 0
        self.brank_length = 0
        self.detail_length = 0
        self.ws['A1'] = "brank"
        self.ws['B1'] = "name"
        self.ws['C1'] = "price"
        self.ws['D1'] = "title"
        self.ws['E1'] = "img"
        self.ws['F1'] = "detail"

    def brank_into(self):
        self.driver.get("https://category.vip.com/suggest.php?keyword=%E5%8F%A3%E7%BA%A2&ff=235")
        self.driver.find_element_by_class_name("c-filter-group-button-text").click()
        ul_list = self.driver.find_elements_by_xpath(
            '//ul[@class="c-filter-data-list  c-filter-brand-list J-brand-filter-data-list J-filter-data-list"]/li[@class="c-filter-data-item  J-filter-data-item"]/a')
        name_list = self.driver.find_elements_by_xpath(
            '//ul[@class="c-filter-data-list  c-filter-brand-list J-brand-filter-data-list J-filter-data-list"]/li[@class="c-filter-data-item  J-filter-data-item"]/a/img')
        self.item["ul_list"] = [i.get_attribute("href") for i in ul_list]
        self.item["name_list"] = [i.get_attribute("alt") for i in name_list]

        return self.item["ul_list"]

    def update_length(self, litem, ws):
        title_length = len(self.litem["title"])
        name_length = len(self.litem["name"][0])
        img_length = len(self.litem["img"][0])
        brank_length = len(self.litem["brank"])
        detail_length = len(
            self.litem["detail"][0] + self.litem["detail"][1] + self.litem["detail"][2] + self.litem["detail"][3])
        if brank_length > self.brank_length:
            self.brank_length = brank_length
        if title_length > self.title_length:
            self.title_length = title_length
        if name_length > self.name_length:
            self.name_length = name_length
        if img_length > self.img_length:
            self.img_length = img_length
        if detail_length > self.detail_length:
            self.detail_length = detail_length
        ws.column_dimensions['A'].width = self.brank_length
        ws.column_dimensions['D'].width = self.title_length
        ws.column_dimensions['B'].width = self.name_length
        ws.column_dimensions['E'].width = self.img_length
        ws.column_dimensions['F'].width = self.detail_length

    def lipstick_get(self, ullist):
        for i in ullist[:2]:

            # print(i)
            self.driver.get(i)
            self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
            hr_list = self.driver.find_elements_by_xpath("//div[@class='goods-slide']/div/a")
            hr_list = [i.get_attribute("href") for i in hr_list]
            # print(litem)
            for hr in hr_list:
                self.driver.get(hr)
                self.litem["href"] = hr
                self.litem["brank"] = self.item["name_list"][ullist.index(i)]
                name_list = self.driver.find_elements_by_xpath(
                    "//div[@class='pi-title-box']/div/p[@class='pib-title-detail']")
                self.litem["name"] = [i.get_attribute("title") for i in name_list]
                price_list = self.driver.find_elements_by_xpath(
                    "//div[@class='sp-info']/span[@class='sp-price']") if self.driver.find_elements_by_xpath(
                    "//div[@class='sp-info']/span[@class='sp-price']") == "" else self.driver.find_elements_by_xpath(
                    "//div[@class='pb-vipPrice']/em")
                self.litem["price"] = [i.text for i in price_list]
                try:
                    title_list = self.driver.find_elements_by_xpath(
                        "//div[@class='pi-title-box']/div/p/span[@class='goods-description-title']")
                    print(title_list)
                    self.litem["title"] = [i.text for i in title_list]
                except:
                    self.litem["title"]="Null"

                img_list = self.driver.find_elements_by_xpath("//div[@class='zoomWindow']//div/img")
                self.litem["img"] = [i.get_attribute("src") for i in img_list]
                detail_list = self.driver.find_elements_by_xpath("//table[@class='dc-table fst']/tbody/tr")
                self.litem["detail"] = [i.text for i in detail_list]
                pprint(self.litem)
                self.update_length(self.litem, self.ws)
                self.ws.append(
                    [
                        self.litem["brank"],
                        self.litem["name"][0],
                        self.litem["price"][0],
                        self.litem["title"][0],
                        self.litem["img"][0],
                        self.litem["detail"][0] + self.litem["detail"][1] + self.litem["detail"][2] +
                        self.litem["detail"][3]
                        # str(self.litem["brank"][0]),
                        # str(self.litem["name"][0]),
                        # str(self.litem["price"][0]),
                        # str(self.litem["title"]),
                        # str(self.litem["img"][0]),
                        # str(self.litem["detail"])
                    ])

                self.wb.save("唯品会.xlsx")

    def __del__(self):
        self.driver.close()

    def run(self):
        ullist = self.brank_into()
        self.lipstick_get(ullist)
        # self.driver.close()


if __name__ == '__main__':
    a = lipstick()
    a.run()
