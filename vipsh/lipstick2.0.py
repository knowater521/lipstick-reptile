import collections
from pprint import pprint
from openpyxl import Workbook
from selenium import webdriver


class lipstick(object):
    def __init__(self):
        self.driver = webdriver.Chrome(r'D:\pycharm\chromedriver_win32\chromedriver.exe')
        self.item = {}
        self.litem = {}
        self.wb = Workbook()  # 得到一个sheet 表
        self.ws = self.wb.active  # 获取第一个sheet
        self.title_length = 0
        self.name_length = 0
        self.img_length = 0
        self.brank_length = 0
        self.detail_length = 0
        self.ws['A1'] = "brank"
        self.ws['B1'] = "name"
        self.ws['C1'] = "price"
        self.ws['D1'] = "title"
        self.ws['E1'] = "img"
        self.ws['F1'] = "detail"

    def brank_into(self):
        self.driver.get("https://category.vip.com/suggest.php?keyword=%E5%8F%A3%E7%BA%A2&ff=235")
        self.driver.find_element_by_class_name("c-filter-group-button-text").click()
        ul_list = self.driver.find_elements_by_xpath(
            '//ul[@class="c-filter-data-list  c-filter-brand-list J-brand-filter-data-list J-filter-data-list"]/li[@class="c-filter-data-item  J-filter-data-item"]/a')
        name_list = self.driver.find_elements_by_xpath(
            '//ul[@class="c-filter-data-list  c-filter-brand-list J-brand-filter-data-list J-filter-data-list"]/li[@class="c-filter-data-item  J-filter-data-item"]/a/img')
        self.item["ul_list"] = [i.get_attribute("href") for i in ul_list]
        self.item["name_list"] = [i.get_attribute("alt") for i in name_list]
        return self.item["ul_list"]

    def next_page(self):
        next_page_list = []
        while True:
            next_page_link = self.driver.find_element_by_xpath(
                './/div[@id="J_pagingCt"]/span[position()>1]/following-sibling::a[1]').get_attribute('href')
            if next_page_link is not None:
                self.driver.get(next_page_link)
                self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
                hr_list = self.driver.find_elements_by_xpath("//div[@class='goods-slide']/div/a")
                hr_list = [i.get_attribute("href") for i in hr_list]
                next_page_list.append(hr_list)

            else:
                break
        return next_page_list

    def update_length(self, litem, ws):
        title_length = len(self.litem["title"])
        name_length = len(self.litem["name"])
        img_length = len(self.litem["img"])
        brank_length = len(self.litem["brank"])
        detail_length = len(self.litem["detail"])

        if brank_length > self.brank_length:
            self.brank_length = brank_length
        if title_length > self.title_length:
            self.title_length = title_length
        if name_length > self.name_length:
            self.name_length = name_length
        if img_length > self.img_length:
            self.img_length = img_length
        if detail_length > self.detail_length:
            self.detail_length = detail_length
        ws.column_dimensions['A'].width = self.brank_length
        ws.column_dimensions['D'].width = self.title_length
        ws.column_dimensions['B'].width = self.name_length
        ws.column_dimensions['E'].width = self.img_length
        ws.column_dimensions['F'].width = self.detail_length

    def lipstick_get(self, ullist):
        for i in ullist:
            # print(i)
            self.driver.get(i)
            self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
            hr_list = self.driver.find_elements_by_xpath("//div[@class='goods-slide']/div/a")
            hr_list = [i.get_attribute("href") for i in hr_list]
            try:
                next_page_list2 = self.next_page()
                a = next_page_list2.append(hr_list)
            except:
                a = hr_list
            for hr in a:
                self.driver.get(hr)
                self.litem["href"] = hr
                self.litem["brank"] = self.item["name_list"][ullist.index(i)]
                self.litem["name"] = self.driver.find_element_by_xpath(
                    "//div[@class='pi-title-box']/div/p[@class='pib-title-detail']").get_attribute("title")
                try:
                    self.litem["price"] = self.driver.find_element_by_xpath(
                        "//div[@class='sp-info']/span[@class='sp-price']").text
                except:
                    self.litem["price"] = self.driver.find_element_by_xpath("//div[@class='pb-vipPrice']/em").text
                try:
                    self.litem["title"] = self.driver.find_element_by_xpath(
                        "//div[@class='pi-title-box']/div/p/span[@class='goods-description-title']").text

                except:
                    self.litem["title"] = "Null"
                try:
                    self.litem["img"] = self.driver.find_element_by_xpath(
                        "//div[@class='zoomWindow']//div/img").get_attribute("src")
                except:
                    self.litem["img"] = "Null"
                self.litem["detail"] = self.driver.find_element_by_xpath("//table[@class='dc-table fst']").text
                pprint(self.litem)
                self.update_length(self.litem, self.ws)
                self.ws.append(
                    [
                        self.litem["brank"],
                        self.litem["name"],
                        self.litem["price"],
                        self.litem["title"],
                        self.litem["img"],
                        self.litem["detail"]

                    ])

                self.wb.save("唯品会.xlsx")

    def __del__(self):
        self.driver.close()

    def run(self):
        ullist = self.brank_into()
        self.lipstick_get(ullist)


if __name__ == '__main__':
    a = lipstick()
    a.run()
